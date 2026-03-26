"""
astar_debug_runner.py
=====================
Instruments for every A* node expansion to capture:
  - g(n)  : accumulated cost so far
  - h_A   : Component A — aggregate energy balance heuristic
  - h_B   : Component B — per-stream temperature obligation heuristic
  - h_C   : Component C — pinch composite curve heuristic
  - h(n)  : max(h_A, h_B, h_C) — final admissible heuristic
  - f(n)  : g(n) + h(n)
  - hot/cold remaining loads
  - action that led to this node

Writes:
  1. astar_debug_log.txt  — human-readable full trace
  2. astar_nodes.xlsx     — spreadsheet with one row per node

Run: python astar_debug_runner.py  (from inside HENS-Astar/)
"""

import os, sys, heapq, time
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

sys.path.insert(0, os.path.dirname(__file__))

from state       import HENSState, HotStream, ColdStream, make_initial_state
from actions     import get_successors
from constraints import current_hot_temp, current_cold_temp
from cost        import STEAM_COST_PER_KW_YEAR, COOLING_COST_PER_KW_YEAR

# ── stream conversion helpers ────────────────────────────────────────────────
def _F(f): return (f - 32) * 5 / 9
def _W(w): return w * 0.000293071

HOT_STREAMS = [
    HotStream("H1", T_in=_F(320), T_out=_F(200), FCp=_W(16670)),
    HotStream("H2", T_in=_F(480), T_out=_F(280), FCp=_W(20000)),
    HotStream("H3", T_in=_F(440), T_out=_F(150), FCp=_W(28000)),
    HotStream("H4", T_in=_F(520), T_out=_F(300), FCp=_W(23800)),
    HotStream("H5", T_in=_F(390), T_out=_F(150), FCp=_W(33600)),
]
COLD_STREAMS = [
    ColdStream("C1", T_in=_F(140), T_out=_F(320), FCp=_W(14450)),
    ColdStream("C2", T_in=_F(240), T_out=_F(431), FCp=_W(11530)),
    ColdStream("C3", T_in=_F(100), T_out=_F(430), FCp=_W(16000)),
    ColdStream("C4", T_in=_F(180), T_out=_F(350), FCp=_W(32760)),
    ColdStream("C5", T_in=_F(200), T_out=_F(400), FCp=_W(26350)),
]
DELTA_T_MIN = 11.1
HOT_DICT  = {h.sid: h for h in HOT_STREAMS}
COLD_DICT = {c.sid: c for c in COLD_STREAMS}

TOLERANCE = 0.5


# ── heuristic decomposition ──────────────────────────────────────────────────
def _build_composite(segments, ascending=False):
    if not segments:
        return [(0.0, 0.0)]
    temps = set()
    for T_h, T_l, _ in segments:
        temps.add(T_h); temps.add(T_l)
    ts = sorted(temps, reverse=True)
    pts, H = [], 0.0
    pts.append((H, ts[0]))
    for i in range(len(ts) - 1):
        T_top, T_bot = ts[i], ts[i + 1]
        fcp = sum(f for (th, tl, f) in segments
                  if th >= T_top - 1e-9 and tl <= T_bot + 1e-9)
        H += fcp * (T_top - T_bot)
        pts.append((H, T_bot))
    return list(reversed(pts)) if ascending else pts


def _problem_table_algorithm(hot_segs, cold_segs_shifted):
    """
    Linnhoff & Hindmarsh (1983) Problem Table Algorithm.
    hot_segs: [(T_hi, T_lo, FCp), ...]  — actual temperatures
    cold_segs_shifted: [(T_hi+dTmin, T_lo+dTmin, FCp), ...]  — shifted up by dTmin
    Returns (QHmin_kW, QCmin_kW).
    """
    all_segs = hot_segs + cold_segs_shifted
    if not all_segs:
        return 0.0, 0.0
    temps = set()
    for T_hi, T_lo, _ in all_segs:
        temps.add(T_hi); temps.add(T_lo)
    ts = sorted(temps, reverse=True)
    if len(ts) < 2:
        return 0.0, 0.0
    residuals = [0.0]
    for i in range(len(ts) - 1):
        T_top, T_bot = ts[i], ts[i + 1]
        hot_fcp  = sum(f for (th, tl, f) in hot_segs
                       if th >= T_top - 1e-9 and tl <= T_bot + 1e-9)
        cold_fcp = sum(f for (th, tl, f) in cold_segs_shifted
                       if th >= T_top - 1e-9 and tl <= T_bot + 1e-9)
        surplus  = (hot_fcp - cold_fcp) * (T_top - T_bot)
        residuals.append(residuals[-1] + surplus)
    QH_min = max(0.0, -min(residuals))
    QC_min = max(0.0, residuals[-1] + QH_min)
    return QH_min, QC_min


def heuristic_full(state, hot_streams, cold_streams, delta_T_min):
    """
    Returns (h_total, h_A, h_B, h_C,
             heat_A, heat_B, heat_C,
             cool_A, cool_B, cool_C,
             heating_util_kW, cooling_util_kW)
    """
    total_hot  = sum(max(0.0, v) for v in state.hot_remaining.values()  if v > TOLERANCE)
    total_cold = sum(max(0.0, v) for v in state.cold_remaining.values() if v > TOLERANCE)

    # Component A
    heat_A = max(0.0, total_cold - total_hot)
    cool_A = max(0.0, total_hot  - total_cold)

    heat_B = cool_B = 0.0
    heat_C = cool_C = 0.0

    if hot_streams and cold_streams:
        max_hot_T = max(
            (current_hot_temp(hot_streams[h], r)
             for h, r in state.hot_remaining.items() if r > TOLERANCE),
            default=0.0,
        )
        min_cold_T = min(
            (current_cold_temp(cold_streams[c], r)
             for c, r in state.cold_remaining.items() if r > TOLERANCE),
            default=9999.0,
        )

        # Component B
        for c_id, c_rem in state.cold_remaining.items():
            if c_rem <= TOLERANCE: continue
            cold = cold_streams[c_id]
            oblig_T = max_hot_T - delta_T_min
            if cold.T_out > oblig_T:
                T_curr = current_cold_temp(cold, c_rem)
                T_bot  = max(T_curr, oblig_T)
                if cold.T_out > T_bot:
                    heat_B += min(cold.FCp * (cold.T_out - T_bot), c_rem)

        for h_id, h_rem in state.hot_remaining.items():
            if h_rem <= TOLERANCE: continue
            hot = hot_streams[h_id]
            oblig_T = min_cold_T + delta_T_min
            if hot.T_out < oblig_T:
                T_curr = current_hot_temp(hot, h_rem)
                T_top  = min(T_curr, oblig_T)
                if T_top > hot.T_out:
                    cool_B += min(hot.FCp * (T_top - hot.T_out), h_rem)

        # Component C: Problem Table Algorithm (Linnhoff & Hindmarsh, 1983)
        # FIX: old composite-curve gap was always <=0 (hot total>>cold total means
        # hot H always dominates cold H when both independently start at 0).
        # PTA correctly unifies breakpoints and cascades surplus.
        try:
            hot_segs = [
                (current_hot_temp(hot_streams[h], r), hot_streams[h].T_out, hot_streams[h].FCp)
                for h, r in state.hot_remaining.items()
                if r > TOLERANCE and current_hot_temp(hot_streams[h], r) > hot_streams[h].T_out + 1e-9
            ]
            cold_segs_shifted = [
                (cold_streams[c].T_out  + delta_T_min,
                 current_cold_temp(cold_streams[c], r) + delta_T_min,
                 cold_streams[c].FCp)
                for c, r in state.cold_remaining.items()
                if r > TOLERANCE and cold_streams[c].T_out > current_cold_temp(cold_streams[c], r) + 1e-9
            ]
            if hot_segs or cold_segs_shifted:
                heat_C, cool_C = _problem_table_algorithm(hot_segs, cold_segs_shifted)
        except Exception:
            heat_C = cool_C = 0.0

    heating_kW = max(heat_A, heat_B, heat_C)
    cooling_kW = max(cool_A, cool_B, cool_C)

    h_A = heat_A * STEAM_COST_PER_KW_YEAR + cool_A * COOLING_COST_PER_KW_YEAR
    h_B = heat_B * STEAM_COST_PER_KW_YEAR + cool_B * COOLING_COST_PER_KW_YEAR
    h_C = heat_C * STEAM_COST_PER_KW_YEAR + cool_C * COOLING_COST_PER_KW_YEAR
    h   = max(h_A, h_B, h_C)

    return (h, h_A, h_B, h_C,
            heat_A, heat_B, heat_C,
            cool_A, cool_B, cool_C,
            heating_kW, cooling_kW)


# ── node record ──────────────────────────────────────────────────────────────
@dataclass
class NodeRecord:
    node_id:        int
    parent_id:      int
    tree_level:     int
    action:         str
    g_cost:         float
    h_total:        float
    f_cost:         float
    h_A:            float
    h_B:            float
    h_C:            float
    heat_A_kW:      float
    heat_B_kW:      float
    heat_C_kW:      float
    cool_A_kW:      float
    cool_B_kW:      float
    cool_C_kW:      float
    heating_util_kW: float
    cooling_util_kW: float
    hot_remaining:  str
    cold_remaining: str
    winning_component: str
    num_exchangers: int
    status:         str    # EXPANDED / GOAL / PRUNED


# ── instrumented A* ──────────────────────────────────────────────────────────
@dataclass(order=True)
class PQItem:
    f_value:    float
    tree_level: int
    counter:    int
    node_id:    int = field(compare=False)
    state: HENSState = field(compare=False)


def run_instrumented_astar(max_nodes=100_000):
    t0 = time.time()
    records: List[NodeRecord] = []
    node_id_map: Dict[int, int] = {}   # hash(state) -> node_id

    start  = make_initial_state(list(HOT_DICT.values()), list(COLD_DICT.values()))
    h_info = heuristic_full(start, HOT_DICT, COLD_DICT, DELTA_T_MIN)
    h0     = h_info[0]

    root_id = 0
    node_id_map[hash(start)] = root_id

    def hot_rem_str(s): return " | ".join(f"{k}:{v:.1f}" for k, v in s.hot_remaining.items())
    def cold_rem_str(s): return " | ".join(f"{k}:{v:.1f}" for k, v in s.cold_remaining.items())

    # Record root
    (h, hA, hB, hC, heatA, heatB, heatC, coolA, coolB, coolC, hkW, ckW) = h_info
    win = ["A","B","C"][[hA,hB,hC].index(max(hA,hB,hC))]
    records.append(NodeRecord(
        node_id=0, parent_id=-1, tree_level=0, action="ROOT",
        g_cost=0, h_total=h, f_cost=h,
        h_A=hA, h_B=hB, h_C=hC,
        heat_A_kW=heatA, heat_B_kW=heatB, heat_C_kW=heatC,
        cool_A_kW=coolA, cool_B_kW=coolB, cool_C_kW=coolC,
        heating_util_kW=hkW, cooling_util_kW=ckW,
        hot_remaining=hot_rem_str(start), cold_remaining=cold_rem_str(start),
        winning_component=win, num_exchangers=0, status="EXPANDED"
    ))

    frontier = []
    heapq.heappush(frontier, PQItem(h0, 0, 0, root_id, start))

    visited = set()
    counter = 1
    nid_counter = 1
    nodes_expanded = 0
    goal_state = None

    while frontier:
        item  = heapq.heappop(frontier)
        state = item.state
        nid   = item.node_id

        key = hash(state)
        if key in visited:
            records[nid].status = "PRUNED"
            continue
        visited.add(key)

        records[nid].status = "EXPANDED"
        nodes_expanded += 1

        if state.is_goal():
            records[nid].status = "GOAL"
            goal_state = state
            break

        if nodes_expanded >= max_nodes:
            break

        successors = get_successors(state, HOT_DICT, COLD_DICT, DELTA_T_MIN)

        for succ in successors:
            sh = hash(succ)
            if sh in visited:
                continue

            hi = heuristic_full(succ, HOT_DICT, COLD_DICT, DELTA_T_MIN)
            (h, hA, hB, hC, heatA, heatB, heatC, coolA, coolB, coolC, hkW, ckW) = hi
            f  = succ.g_cost + h

            # determine action label
            if len(succ.matches) > len(state.matches):
                m = succ.matches[-1]
                act = f"MATCH({m.hot_id},{m.cold_id}) {m.duty:.0f}kW"
            elif len(succ.heaters) > len(state.heaters):
                ht = succ.heaters[-1]
                act = f"HEATER({ht.cold_id}) {ht.duty:.0f}kW"
            elif len(succ.coolers) > len(state.coolers):
                ct = succ.coolers[-1]
                act = f"COOLER({ct.hot_id}) {ct.duty:.0f}kW"
            else:
                act = "UNKNOWN"

            win = ["A","B","C"][[hA,hB,hC].index(max(hA,hB,hC))]
            new_nid = nid_counter
            nid_counter += 1
            node_id_map[sh] = new_nid

            records.append(NodeRecord(
                node_id=new_nid, parent_id=nid, tree_level=succ.tree_level,
                action=act,
                g_cost=succ.g_cost, h_total=h, f_cost=f,
                h_A=hA, h_B=hB, h_C=hC,
                heat_A_kW=heatA, heat_B_kW=heatB, heat_C_kW=heatC,
                cool_A_kW=coolA, cool_B_kW=coolB, cool_C_kW=coolC,
                heating_util_kW=hkW, cooling_util_kW=ckW,
                hot_remaining=hot_rem_str(succ), cold_remaining=cold_rem_str(succ),
                winning_component=win,
                num_exchangers=succ.num_exchangers(),
                status="QUEUED"
            ))

            heapq.heappush(frontier, PQItem(f, succ.tree_level, counter, new_nid, succ))
            counter += 1

    elapsed = time.time() - t0
    return records, goal_state, nodes_expanded, elapsed


# ── TXT writer ───────────────────────────────────────────────────────────────
def write_txt(records, goal_state, nodes_expanded, elapsed, path):
    W = 110
    with open(path, "w", encoding="utf-8") as f:
        def w(s=""): f.write(s + "\n")

        w("=" * W)
        w("  HENS A* — FULL NODE-BY-NODE HEURISTIC & F-COST TRACE")
        w("  Problem : Pho & Lapidus (1973) 10SP1 — 5H x 5C streams")
        w("  dT_min  : 11.1 °C  |  Steam: $160/kW·yr  |  CW: $60/kW·yr")
        w("=" * W)

        w()
        w("HEURISTIC COMPONENTS (all admissible lower bounds on remaining TAC):")
        w("  h_A  = Aggregate energy balance  — global hot/cold surplus or deficit")
        w("  h_B  = Per-stream temperature obligation — portions forced to utility")
        w("         by dT_min limits regardless of any match")
        w("  h_C  = Pinch composite curve — QHmin/QCmin from shifted composite curves")
        w("  h(n) = max(h_A, h_B, h_C) converted to $/yr via utility cost rates")
        w("  f(n) = g(n) + h(n)   [g = accumulated TAC cost placed so far]")
        w()
        w("ADMISSIBILITY PROOF SKETCH:")
        w("  All three components ignore future exchanger capital costs entirely,")
        w("  and use minimum utility rates. Therefore h(n) ≤ h*(n) always. ✓")
        w()
        w("=" * W)
        w()

        # Stream data
        w("STREAM DATA")
        w("-" * 70)
        w(f"  {'ID':<4} {'Type':<5} {'T_in':>8} {'T_out':>8} {'FCp':>9} {'Q_total':>11}")
        w("-" * 70)
        from state import HotStream, ColdStream
        for h in HOT_STREAMS:
            w(f"  {h.sid:<4} {'Hot':<5} {h.T_in:>8.2f} {h.T_out:>8.2f} {h.FCp:>9.4f} {h.Q_total:>10.1f} kW")
        for c in COLD_STREAMS:
            w(f"  {c.sid:<4} {'Cold':<5} {c.T_in:>8.2f} {c.T_out:>8.2f} {c.FCp:>9.4f} {c.Q_total:>10.1f} kW")
        w()

        w("=" * W)
        w("  NODE-BY-NODE TRACE")
        w("=" * W)

        for rec in records:
            w()
            w(f"{'─'*W}")
            status_tag = f"[{rec.status}]"
            w(f"  NODE {rec.node_id:>5}  {status_tag:<12}  Level={rec.tree_level}  "
              f"HXs={rec.num_exchangers}  Action: {rec.action}")
            w(f"{'─'*W}")
            w(f"  ┌─ COST BREAKDOWN")
            w(f"  │   g(n)  = {rec.g_cost:>14,.2f}  $/yr   (accumulated TAC to this node)")
            w(f"  │   h(n)  = {rec.h_total:>14,.2f}  $/yr   (= max of A, B, C below)")
            w(f"  │   f(n)  = {rec.f_cost:>14,.2f}  $/yr   (priority in open list)")
            w(f"  │")
            w(f"  ├─ HEURISTIC COMPONENTS  (kW values → multiplied by utility $/kW·yr)")
            w(f"  │   Component A — Aggregate Energy Balance")
            w(f"  │     heat_A = {rec.heat_A_kW:>10.2f} kW   →  h_A(heat) = {rec.heat_A_kW*160:>12,.2f} $/yr")
            w(f"  │     cool_A = {rec.cool_A_kW:>10.2f} kW   →  h_A(cool) = {rec.cool_A_kW*60:>12,.2f} $/yr")
            w(f"  │     h_A   = {rec.h_A:>14,.2f}  $/yr")
            w(f"  │")
            w(f"  │   Component B — Per-Stream Temperature Obligation")
            w(f"  │     heat_B = {rec.heat_B_kW:>10.2f} kW   →  h_B(heat) = {rec.heat_B_kW*160:>12,.2f} $/yr")
            w(f"  │     cool_B = {rec.cool_B_kW:>10.2f} kW   →  h_B(cool) = {rec.cool_B_kW*60:>12,.2f} $/yr")
            w(f"  │     h_B   = {rec.h_B:>14,.2f}  $/yr")
            w(f"  │")
            w(f"  │   Component C — Pinch Composite Curve")
            w(f"  │     heat_C = {rec.heat_C_kW:>10.2f} kW   →  h_C(heat) = {rec.heat_C_kW*160:>12,.2f} $/yr")
            w(f"  │     cool_C = {rec.cool_C_kW:>10.2f} kW   →  h_C(cool) = {rec.cool_C_kW*60:>12,.2f} $/yr")
            w(f"  │     h_C   = {rec.h_C:>14,.2f}  $/yr")
            w(f"  │")
            win_lbl = {"A":"Component A WINS","B":"Component B WINS","C":"Component C WINS"}
            w(f"  │   ★ WINNING: {win_lbl.get(rec.winning_component, '?')}")
            w(f"  │   Final utility estimate: Heating = {rec.heating_util_kW:.2f} kW, "
              f"Cooling = {rec.cooling_util_kW:.2f} kW")
            w(f"  │")
            w(f"  └─ REMAINING LOADS")
            w(f"       HOT : {rec.hot_remaining}")
            w(f"       COLD: {rec.cold_remaining}")

        w()
        w("=" * W)
        w("  SEARCH SUMMARY")
        w("=" * W)
        w(f"  Total nodes generated : {len(records):,}")
        w(f"  Nodes expanded        : {nodes_expanded:,}")
        w(f"  Elapsed time          : {elapsed:.3f} s")
        if goal_state:
            w(f"  GOAL FOUND at level   : {goal_state.tree_level}")
            w(f"  Optimal TAC           : ${goal_state.g_cost:,.2f} /yr")
            w(f"  Exchangers placed     : {goal_state.num_exchangers()}")
            w(f"  Utility units         : {goal_state.num_utilities()}")
            w()
            w("  WINNING NETWORK:")
            w(str(goal_state.matrix))
            w()
            w("  MATCHES:")
            for m in goal_state.matches:
                w(f"    {m}")
            w("  HEATERS:")
            for h in goal_state.heaters:
                w(f"    {h}")
            w("  COOLERS:")
            for c in goal_state.coolers:
                w(f"    {c}")
        else:
            w("  No goal found within node cap.")

        w()
        w("=" * W)
        w("  INVESTIGATION: ARE ALL 3 HEURISTIC COMPONENTS NEEDED?")
        w("=" * W)

        # count which component won at each expanded node
        wins = {"A": 0, "B": 0, "C": 0}
        h_only_A = h_only_B = h_only_C = 0
        tighter_B_over_A = tighter_C_over_A = tighter_C_over_B = 0
        expanded = [r for r in records if r.status in ("EXPANDED","GOAL")]

        for r in expanded:
            wins[r.winning_component] += 1
            if r.h_B > r.h_A: tighter_B_over_A += 1
            if r.h_C > r.h_A: tighter_C_over_A += 1
            if r.h_C > r.h_B: tighter_C_over_B += 1

        n_exp = len(expanded)
        w(f"  Expanded nodes analysed: {n_exp}")
        w()
        w("  How often each component supplied the TIGHTEST (winning) bound:")
        for comp in ["A","B","C"]:
            pct = 100*wins[comp]/n_exp if n_exp else 0
            bar = "█" * int(pct / 2)
            w(f"    Component {comp}: {wins[comp]:>5} nodes  ({pct:5.1f}%)  {bar}")
        w()
        w("  How often each component was STRICTLY TIGHTER than another:")
        w(f"    B > A : {tighter_B_over_A:>5} / {n_exp}  ({100*tighter_B_over_A/n_exp if n_exp else 0:.1f}%)")
        w(f"    C > A : {tighter_C_over_A:>5} / {n_exp}  ({100*tighter_C_over_A/n_exp if n_exp else 0:.1f}%)")
        w(f"    C > B : {tighter_C_over_B:>5} / {n_exp}  ({100*tighter_C_over_B/n_exp if n_exp else 0:.1f}%)")
        w()
        w("  VERDICT:")
        for comp, name in [("A","Aggregate balance"),("B","Temp. obligation"),("C","Pinch composite")]:
            pct = 100*wins[comp]/n_exp if n_exp else 0
            need = "NEEDED" if wins[comp] > 0 else "REDUNDANT (never wins)"
            w(f"    Component {comp} ({name:<22}): {need}  — wins on {pct:.1f}% of expanded nodes")
        w()
        w("  If any component never wins, it adds computation without tightening h(n).")
        w("  If a component wins frequently, removing it would weaken the heuristic,")
        w("  potentially expanding many more nodes and slowing the search.")
        w("=" * W)


# ── XLSX writer ──────────────────────────────────────────────────────────────
def write_xlsx(records, goal_state, nodes_expanded, elapsed, path):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                 numbers as xlnums)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Node trace ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Node Trace"

    HDR_FILL   = PatternFill("solid", fgColor="1F3864")
    HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    BOLD_FONT  = Font(bold=True, name="Arial", size=10)
    NORM_FONT  = Font(name="Arial", size=10)
    GOAL_FILL  = PatternFill("solid", fgColor="C6EFCE")
    PRUNED_FILL = PatternFill("solid", fgColor="FCE4D6")
    EXPAND_FILL = PatternFill("solid", fgColor="DDEBF7")
    WIN_A_FILL = PatternFill("solid", fgColor="FFF2CC")
    WIN_B_FILL = PatternFill("solid", fgColor="E2EFDA")
    WIN_C_FILL = PatternFill("solid", fgColor="FCE4D6")
    CENTER     = Alignment(horizontal="center", vertical="center")
    LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    RIGHT      = Alignment(horizontal="right",  vertical="center")
    thin       = Side(style="thin", color="BFBFBF")
    BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLS = [
        ("Node ID",          8),
        ("Parent ID",        8),
        ("Level",            6),
        ("Status",           9),
        ("Action",          30),
        ("HXs",              5),
        ("g(n) $/yr",       13),
        ("h_A $/yr",        13),
        ("h_B $/yr",        13),
        ("h_C $/yr",        13),
        ("h(n) $/yr",       13),
        ("f(n) $/yr",       13),
        ("Win Comp",         8),
        ("heat_A kW",       10),
        ("heat_B kW",       10),
        ("heat_C kW",       10),
        ("cool_A kW",       10),
        ("cool_B kW",       10),
        ("cool_C kW",       10),
        ("Heat Util kW",    11),
        ("Cool Util kW",    11),
        ("Hot Remaining",   42),
        ("Cold Remaining",  42),
    ]

    # Header row
    ws.row_dimensions[1].height = 28
    for ci, (col_name, col_w) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = col_w

    money_fmt = '#,##0.00'

    for ri, rec in enumerate(records, 2):
        row_data = [
            rec.node_id, rec.parent_id, rec.tree_level, rec.status,
            rec.action, rec.num_exchangers,
            rec.g_cost, rec.h_A, rec.h_B, rec.h_C, rec.h_total, rec.f_cost,
            rec.winning_component,
            rec.heat_A_kW, rec.heat_B_kW, rec.heat_C_kW,
            rec.cool_A_kW, rec.cool_B_kW, rec.cool_C_kW,
            rec.heating_util_kW, rec.cooling_util_kW,
            rec.hot_remaining, rec.cold_remaining,
        ]

        if rec.status == "GOAL":
            row_fill = GOAL_FILL
        elif rec.status == "PRUNED":
            row_fill = PRUNED_FILL
        elif rec.winning_component == "A":
            row_fill = WIN_A_FILL
        elif rec.winning_component == "B":
            row_fill = WIN_B_FILL
        elif rec.winning_component == "C":
            row_fill = WIN_C_FILL
        else:
            row_fill = EXPAND_FILL

        ws.row_dimensions[ri].height = 18
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font   = NORM_FONT
            cell.fill   = row_fill
            cell.border = BORDER
            if ci in (7, 8, 9, 10, 11, 12):   # money columns
                cell.number_format = money_fmt
                cell.alignment = RIGHT
            elif ci in (14,15,16,17,18,19,20,21):   # kW columns
                cell.number_format = '#,##0.00'
                cell.alignment = RIGHT
            elif ci in (22, 23):               # long text
                cell.alignment = LEFT
            else:
                cell.alignment = CENTER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # ── Sheet 2: Component Win Statistics ────────────────────────────────────
    ws2 = wb.create_sheet("Component Analysis")
    expanded = [r for r in records if r.status in ("EXPANDED","GOAL")]
    n_exp = len(expanded)
    wins = {"A": 0, "B": 0, "C": 0}
    tb = {"B>A": 0, "C>A": 0, "C>B": 0}
    for r in expanded:
        wins[r.winning_component] += 1
        if r.h_B > r.h_A: tb["B>A"] += 1
        if r.h_C > r.h_A: tb["C>A"] += 1
        if r.h_C > r.h_B: tb["C>B"] += 1

    def ws2_header(row, text):
        c = ws2.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = CENTER
        ws2.merge_cells(f"A{row}:D{row}")

    def ws2_row(row, a, b, c, d=""):
        for ci, v in enumerate([a, b, c, d], 1):
            cell = ws2.cell(row=row, column=ci, value=v)
            cell.font   = NORM_FONT
            cell.border = BORDER
            cell.alignment = CENTER if ci > 1 else LEFT

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 22

    ws2_header(1, "HEURISTIC COMPONENT ANALYSIS — ARE ALL 3 NEEDED?")
    ws2.row_dimensions[1].height = 24

    ws2.cell(row=2, column=1, value="Component").font = HDR_FONT
    ws2.cell(row=2, column=1).fill = HDR_FILL
    ws2.cell(row=2, column=1).alignment = CENTER
    ws2.cell(row=2, column=2, value="Wins (nodes)").font = HDR_FONT
    ws2.cell(row=2, column=2).fill = HDR_FILL
    ws2.cell(row=2, column=2).alignment = CENTER
    ws2.cell(row=2, column=3, value="Win %").font = HDR_FONT
    ws2.cell(row=2, column=3).fill = HDR_FILL
    ws2.cell(row=2, column=3).alignment = CENTER
    ws2.cell(row=2, column=4, value="Verdict").font = HDR_FONT
    ws2.cell(row=2, column=4).fill = HDR_FILL
    ws2.cell(row=2, column=4).alignment = CENTER
    for ci in range(1, 5):
        ws2.cell(row=2, column=ci).border = BORDER

    comp_names = {
        "A": "A — Aggregate Energy Balance",
        "B": "B — Temp. Obligation (per-stream)",
        "C": "C — Pinch Composite Curve",
    }
    for i, comp in enumerate(["A","B","C"], 3):
        pct = wins[comp]/n_exp if n_exp else 0
        verdict = "NEEDED" if wins[comp] > 0 else "REDUNDANT — never tightest"
        ws2_row(i, comp_names[comp], wins[comp], f"{pct:.1%}", verdict)
        ws2.cell(row=i, column=3).number_format = "0.0%"
        ws2.cell(row=i, column=3).value = pct

    ws2_row(6, "Total expanded nodes", n_exp, "", "")

    ws2_header(8, "STRICTLY TIGHTER COMPARISONS (expanded nodes)")
    ws2.cell(row=9, column=1, value="Comparison").font = HDR_FONT
    ws2.cell(row=9, column=1).fill = HDR_FILL; ws2.cell(row=9, column=1).border = BORDER
    ws2.cell(row=9, column=1).alignment = CENTER
    ws2.cell(row=9, column=2, value="Count").font = HDR_FONT
    ws2.cell(row=9, column=2).fill = HDR_FILL; ws2.cell(row=9, column=2).border = BORDER
    ws2.cell(row=9, column=2).alignment = CENTER
    ws2.cell(row=9, column=3, value="% of expanded").font = HDR_FONT
    ws2.cell(row=9, column=3).fill = HDR_FILL; ws2.cell(row=9, column=3).border = BORDER
    ws2.cell(row=9, column=3).alignment = CENTER
    ws2.cell(row=9, column=4, value="Meaning").font = HDR_FONT
    ws2.cell(row=9, column=4).fill = HDR_FILL; ws2.cell(row=9, column=4).border = BORDER
    ws2.cell(row=9, column=4).alignment = CENTER

    meanings = {
        "B>A": "B tighter than A alone → B adds value",
        "C>A": "C tighter than A alone → C adds value",
        "C>B": "C tighter than B → C adds value over B",
    }
    for i, (key, meaning) in enumerate(meanings.items(), 10):
        pct = tb[key]/n_exp if n_exp else 0
        ws2_row(i, key, tb[key], pct, meaning)
        ws2.cell(row=i, column=3).number_format = "0.0%"
        ws2.cell(row=i, column=3).value = pct

    ws2_header(13, "SEARCH PERFORMANCE SUMMARY")
    perf = [
        ("Nodes generated (total)",  len(records)),
        ("Nodes expanded",           nodes_expanded),
        ("Elapsed time (s)",         round(elapsed, 4)),
    ]
    if goal_state:
        perf += [
            ("Optimal TAC ($/yr)",       round(goal_state.g_cost, 2)),
            ("Solution depth (level)",   goal_state.tree_level),
            ("Process exchangers",       goal_state.num_exchangers()),
            ("Utility units",            goal_state.num_utilities()),
        ]
    for i, (label, val) in enumerate(perf, 14):
        ws2.cell(row=i, column=1, value=label).font = NORM_FONT
        ws2.cell(row=i, column=1).border = BORDER
        ws2.cell(row=i, column=2, value=val).font  = BOLD_FONT
        ws2.cell(row=i, column=2).border = BORDER
        ws2.cell(row=i, column=2).alignment = CENTER

    # ── Sheet 3: Per-level heuristic averages ────────────────────────────────
    ws3 = wb.create_sheet("Level Averages")
    from collections import defaultdict
    lvl = defaultdict(list)
    for r in expanded:
        lvl[r.tree_level].append(r)

    ws3.column_dimensions["A"].width = 8
    for ci, w_ in enumerate([13,13,13,13,13,10,10], 2):
        ws3.column_dimensions[get_column_letter(ci)].width = w_

    hdrs = ["Level","Avg g(n)","Avg h_A","Avg h_B","Avg h_C","Avg h(n)","Avg f(n)","Nodes"]
    for ci, h_ in enumerate(hdrs, 1):
        c = ws3.cell(row=1, column=ci, value=h_)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER

    for ri, level in enumerate(sorted(lvl.keys()), 2):
        recs = lvl[level]
        n = len(recs)
        def avg(attr): return sum(getattr(r, attr) for r in recs) / n
        row_vals = [level, avg("g_cost"), avg("h_A"), avg("h_B"), avg("h_C"),
                    avg("h_total"), avg("f_cost"), n]
        for ci, v in enumerate(row_vals, 1):
            c = ws3.cell(row=ri, column=ci, value=round(v, 2) if isinstance(v, float) else v)
            c.font   = NORM_FONT
            c.border = BORDER
            c.alignment = RIGHT if ci > 1 else CENTER
            if ci in (2,3,4,5,6,7):
                c.number_format = money_fmt

    wb.save(path)
    print(f"  XLSX saved: {path}")


# ── main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Running instrumented A* — logging every node...")
    records, goal_state, nodes_expanded, elapsed = run_instrumented_astar(max_nodes=100_000)
    print(f"  Done. {len(records):,} nodes recorded, {nodes_expanded:,} expanded, {elapsed:.2f}s")

    txt_path  = "/home/claude/astar_debug_log.txt"
    xlsx_path = "/home/claude/astar_nodes.xlsx"

    print("Writing TXT log...")
    write_txt(records, goal_state, nodes_expanded, elapsed, txt_path)
    print(f"  TXT saved: {txt_path}")

    print("Writing XLSX spreadsheet...")
    write_xlsx(records, goal_state, nodes_expanded, elapsed, xlsx_path)

    print("All done.")
